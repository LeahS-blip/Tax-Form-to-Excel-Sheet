"""
Write extracted fields to a formatted .xlsx workbook.

Layout:
  - One sheet per processed file ("W-2 — acme.pdf", etc.) with Field / Value /
    Confidence / Source columns.
  - A "Summary" sheet with one row per file and the headline numbers, so a
    folder of W-2s becomes a single comparable table.

Confidence below 0.5 is flagged amber, missing values red, so a human can scan
for what needs checking — important for anything tax-related.
"""

from __future__ import annotations
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .parsers import FieldResult

_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_LOW_FILL = PatternFill("solid", fgColor="FFF2CC")     # amber: low confidence
_MISSING_FILL = PatternFill("solid", fgColor="FBE4E4")  # red: no value
_TITLE_FONT = Font(bold=True, size=14, color="1F3A5F")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY_FMT = "#,##0.00"


@dataclass
class FileResult:
    filename: str
    form_type: str
    source: str            # native / ocr
    engine: str            # positional / llm
    fields: dict[str, FieldResult]


def _safe_title(s: str) -> str:
    for ch in r'[]:*?/\\':
        s = s.replace(ch, "-")
    return s[:31]


def _style_header(ws, row, headers):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER


def _write_detail_sheet(wb: Workbook, fr: FileResult):
    ws = wb.create_sheet(_safe_title(f"{fr.form_type} — {fr.filename}"))
    ws["A1"] = f"{fr.form_type}  ·  {fr.filename}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"text source: {fr.source}   ·   engine: {fr.engine}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    _style_header(ws, 4, ["Field", "Value", "Confidence", "Note"])
    r = 5
    for res in fr.fields.values():
        ws.cell(r, 1, res.label).border = _BORDER
        vcell = ws.cell(r, 2)
        if isinstance(res.value, (int, float)):
            vcell.value = res.value
            vcell.number_format = _MONEY_FMT
        else:
            vcell.value = res.value if res.value not in (None, "") else None
        vcell.border = _BORDER
        ccell = ws.cell(r, 3, round(res.confidence, 2))
        ccell.border = _BORDER
        note = ws.cell(r, 4)
        note.border = _BORDER
        if res.value in (None, ""):
            for col in (1, 2, 3, 4):
                ws.cell(r, col).fill = _MISSING_FILL
            note.value = "not found — check manually"
        elif res.confidence < 0.5:
            for col in (1, 2, 3, 4):
                ws.cell(r, col).fill = _LOW_FILL
            note.value = "low confidence — verify"
        r += 1

    widths = [38, 22, 12, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


# Headline fields shown on the Summary sheet per form type.
_SUMMARY_FIELDS = {
    "W-2": ["employee_name", "employer_name", "box1_wages",
            "box2_fed_withheld", "box3_ss_wages", "box5_medicare_wages",
            "box16_state_wages", "box17_state_tax"],
    "1040": ["agi", "taxable_income", "total_tax", "fed_withholding",
             "total_payments", "overpayment", "amount_owed"],
    "Schedule C": ["proprietor_name", "business_name", "gross_receipts",
                   "gross_income", "total_expenses", "net_profit"],
    "Schedule E": ["total_rents", "total_royalties", "total_expenses",
                   "total_income_loss"],
    "Schedule K-1": ["entity_name", "recipient_name",
                     "ordinary_business_income", "net_rental_real_estate",
                     "interest_income", "ordinary_dividends",
                     "self_employment_earnings"],
    "1099-NEC": ["payer_name", "recipient_name", "box1_nonemployee_comp",
                 "box4_fed_withheld"],
    "1099-MISC": ["payer_name", "box1_rents", "box2_royalties",
                  "box3_other_income", "box4_fed_withheld"],
    "1099-INT": ["payer_name", "box1_interest_income",
                 "box3_treasury_interest", "box4_fed_withheld"],
    "1099-DIV": ["payer_name", "box1a_ordinary_dividends",
                 "box1b_qualified_dividends", "box2a_capital_gain",
                 "box4_fed_withheld"],
    "1099-R": ["payer_name", "box1_gross_distribution",
               "box2a_taxable_amount", "box4_fed_withheld",
               "box7_distribution_code"],
}


def _write_summary(wb: Workbook, results: list[FileResult]):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Tax form extraction summary"
    ws["A1"].font = _TITLE_FONT

    # Group by form type so headers line up.
    by_type: dict[str, list[FileResult]] = {}
    for fr in results:
        by_type.setdefault(fr.form_type, []).append(fr)

    row = 3
    for ftype, group in by_type.items():
        keys = _SUMMARY_FIELDS.get(ftype, [])
        labels = []
        for k in keys:
            lbl = next((f.label for f in group[0].fields.values()
                        if f.key == k), k)
            labels.append(lbl)
        ws.cell(row, 1, ftype).font = Font(bold=True, size=12, color="1F3A5F")
        row += 1
        _style_header(ws, row, ["File"] + labels)
        row += 1
        for fr in group:
            ws.cell(row, 1, fr.filename).border = _BORDER
            for c, k in enumerate(keys, start=2):
                res = fr.fields.get(k)
                cell = ws.cell(row, c)
                cell.border = _BORDER
                if res and isinstance(res.value, (int, float)):
                    cell.value = res.value
                    cell.number_format = _MONEY_FMT
                elif res:
                    cell.value = res.value
            row += 1
        row += 2

    ws.column_dimensions["A"].width = 26
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 18


def write_workbook(results: list[FileResult], out_path: str):
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    for fr in results:
        _write_detail_sheet(wb, fr)
    _write_summary(wb, results)
    wb.save(out_path)
    return out_path
