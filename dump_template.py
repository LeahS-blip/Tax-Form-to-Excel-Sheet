"""
One-off helper: dump the structure of the Smith_Neal verification template so
the layout can be mirrored in the extractor's Excel writer.

Run it:
    C:\Python314\python.exe "C:\Users\leahs\OneDrive\Documents\PDF to Excel\dump_template.py"

It writes _template_structure.txt next to this script. Nothing is uploaded.
"""
import os
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "Smith_Neal_2025_Tax_Verification Example.xlsx")
OUT = os.path.join(HERE, "_template_structure.txt")


def main():
    if not os.path.exists(SRC):
        raise SystemExit(f"Could not find the template at:\n  {SRC}\n"
                         "Put this script in the same folder as the .xlsx.")

    wb = load_workbook(SRC, data_only=True)      # values
    wbf = load_workbook(SRC, data_only=False)    # formulas + styling

    lines = []
    lines.append(f"FILE: {os.path.basename(SRC)}")
    lines.append(f"SHEETS: {wb.sheetnames}")

    for ws, wsf in zip(wb.worksheets, wbf.worksheets):
        lines.append("")
        lines.append(f"===== SHEET: {ws.title!r}  dims={ws.dimensions}  "
                     f"max_row={ws.max_row} max_col={ws.max_column} =====")
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            lines.append(f"  merged cells: {merged}")
        # Column widths
        widths = {k: round(v.width, 1) for k, v in ws.column_dimensions.items()
                  if v.width}
        if widths:
            lines.append(f"  col widths: {widths}")
        lines.append("  --- rows (value | [formula if any]) ---")
        for row in wsf.iter_rows():
            cells = []
            any_val = False
            for c in row:
                v = c.value
                if v is None:
                    continue
                any_val = True
                coord = c.coordinate
                bold = ""
                try:
                    if c.font and c.font.bold:
                        bold = " <b>"
                except Exception:
                    pass
                cells.append(f"{coord}={str(v)[:60]}{bold}")
            if any_val:
                lines.append("   " + " | ".join(cells))

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Wrote structure to:\n  {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
